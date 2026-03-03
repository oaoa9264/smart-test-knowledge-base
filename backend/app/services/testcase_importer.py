import os
import tempfile
from dataclasses import dataclass
from io import BytesIO
from typing import Any, Dict, Iterable, List, Optional, Tuple


HEADER_KEYWORDS = {
    "title": ["标题", "用例名称", "用例标题", "用例", "title", "case", "name"],
    "steps": ["步骤", "执行步骤", "操作步骤", "steps", "step", "action"],
    "expected_result": ["预期", "预期结果", "期望结果", "expected", "result"],
}


@dataclass
class ParsedTestCase:
    title: str
    steps: str
    expected_result: str
    raw_text: str


def parse_testcases_from_upload(filename: str, file_bytes: bytes) -> List[ParsedTestCase]:
    ext = os.path.splitext((filename or "").lower())[1]
    if ext in [".xlsx", ".xlsm"]:
        return parse_testcases_from_excel(file_bytes)
    if ext == ".xmind":
        return parse_testcases_from_xmind(file_bytes)
    raise ValueError("unsupported file type, only .xlsx/.xlsm/.xmind are allowed")


def parse_testcases_from_excel(file_bytes: bytes) -> List[ParsedTestCase]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover - environment dependency
        raise RuntimeError("openpyxl is required for excel parsing") from exc

    workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    if not rows:
        return []

    header_row_index, column_map = _detect_excel_header(rows)
    start_row = header_row_index + 1 if header_row_index is not None else 0

    parsed_cases: List[ParsedTestCase] = []
    for row in rows[start_row:]:
        title = _safe_cell_text(row, column_map.get("title", 0)).strip()
        steps = _safe_cell_text(row, column_map.get("steps", 1)).strip()
        expected_result = _safe_cell_text(row, column_map.get("expected_result", 2)).strip()

        if not title and not steps and not expected_result:
            continue

        if not title:
            title = (steps or expected_result or "未命名用例")[:80]
        if not steps:
            steps = "-"
        if not expected_result:
            expected_result = "-"

        raw_text = " ".join([part for part in [title, steps, expected_result] if part]).strip()
        parsed_cases.append(
            ParsedTestCase(
                title=title,
                steps=steps,
                expected_result=expected_result,
                raw_text=raw_text,
            )
        )
    return parsed_cases


def parse_testcases_from_xmind(file_bytes: bytes) -> List[ParsedTestCase]:
    try:
        from xmindparser import xmind_to_dict
    except Exception as exc:  # pragma: no cover - environment dependency
        raise RuntimeError("xmindparser is required for xmind parsing") from exc

    with tempfile.NamedTemporaryFile(suffix=".xmind", delete=False) as temp_file:
        temp_file.write(file_bytes)
        temp_path = temp_file.name

    try:
        parsed = xmind_to_dict(temp_path)
    finally:
        try:
            os.remove(temp_path)
        except OSError:
            pass

    if not parsed:
        return []

    cases: List[ParsedTestCase] = []
    for sheet in parsed:
        root = _extract_xmind_root(sheet)
        if not root:
            continue
        leaf_paths = _collect_leaf_paths(root)
        for path in leaf_paths:
            case = _path_to_case(path)
            if case:
                cases.append(case)
    return cases


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return " ".join(text.split())


def _safe_cell_text(row: List[Any], index: int) -> str:
    if index < 0 or index >= len(row):
        return ""
    return _normalize_text(row[index])


def _is_keyword_match(text: str, keywords: Iterable[str]) -> bool:
    normalized = text.lower().replace(" ", "")
    for keyword in keywords:
        key = keyword.lower().replace(" ", "")
        if key and key in normalized:
            return True
    return False


def _detect_excel_header(rows: List[List[Any]]) -> Tuple[Optional[int], Dict[str, int]]:
    best_row_index: Optional[int] = None
    best_score = -1
    best_mapping: Dict[str, int] = {}

    for row_index, row in enumerate(rows[:10]):
        mapping: Dict[str, int] = {}
        for col_index, value in enumerate(row):
            text = _normalize_text(value)
            if not text:
                continue
            for field, keywords in HEADER_KEYWORDS.items():
                if field in mapping:
                    continue
                if _is_keyword_match(text, keywords):
                    mapping[field] = col_index

        score = len(mapping)
        if score > best_score:
            best_score = score
            best_row_index = row_index if score >= 2 else None
            best_mapping = mapping

    if not best_mapping:
        return None, {"title": 0, "steps": 1, "expected_result": 2}

    fallback_mapping = {"title": 0, "steps": 1, "expected_result": 2}
    fallback_mapping.update(best_mapping)
    return best_row_index, fallback_mapping


def _extract_xmind_root(sheet: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    for key in ["topic", "rootTopic", "root_topic"]:
        value = sheet.get(key)
        if isinstance(value, dict):
            return value
    return None


def _extract_topic_title(topic: Dict[str, Any]) -> str:
    for key in ["title", "topicTitle", "text"]:
        value = topic.get(key)
        if isinstance(value, str) and value.strip():
            return value.strip()
    return ""


def _extract_topic_children(topic: Dict[str, Any]) -> List[Dict[str, Any]]:
    children: List[Dict[str, Any]] = []

    topics = topic.get("topics")
    if isinstance(topics, list):
        children.extend([item for item in topics if isinstance(item, dict)])

    children_block = topic.get("children")
    if isinstance(children_block, dict):
        for key in ["attached", "detached"]:
            value = children_block.get(key)
            if isinstance(value, list):
                children.extend([item for item in value if isinstance(item, dict)])
            elif isinstance(value, dict):
                children.append(value)
    elif isinstance(children_block, list):
        children.extend([item for item in children_block if isinstance(item, dict)])

    return children


def _collect_leaf_paths(root: Dict[str, Any]) -> List[List[str]]:
    paths: List[List[str]] = []

    def walk(topic: Dict[str, Any], stack: List[str]) -> None:
        title = _extract_topic_title(topic)
        if title:
            stack = [*stack, title]

        children = _extract_topic_children(topic)
        if not children:
            if stack:
                paths.append(stack)
            return

        for child in children:
            walk(child, stack)

    walk(root, [])
    return paths


def _looks_like_expected_result(text: str) -> bool:
    lowered = text.lower()
    return any(marker in lowered for marker in ["预期", "期望", "expected", "result", "应"])


def _path_to_case(path: List[str]) -> Optional[ParsedTestCase]:
    cleaned = [_normalize_text(item) for item in path if _normalize_text(item)]
    if len(cleaned) <= 1:
        if not cleaned:
            return None
        title = cleaned[0]
        return ParsedTestCase(title=title, steps="-", expected_result="-", raw_text=title)

    branch = cleaned[1:] if len(cleaned) >= 2 else cleaned
    if len(branch) == 1:
        title = branch[0]
        return ParsedTestCase(title=title, steps="-", expected_result="-", raw_text=title)

    module = branch[0]
    title = branch[1] if len(branch) >= 2 else branch[0]
    details = branch[2:] if len(branch) > 2 else []

    step_parts: List[str] = []
    expected_parts: List[str] = []
    for item in details:
        if _looks_like_expected_result(item):
            expected_parts.append(item)
        else:
            step_parts.append(item)

    if not step_parts and details:
        step_parts = details[:-1] if len(details) > 1 else details
    if not expected_parts and details:
        expected_parts = [details[-1]]

    display_title = "{0} - {1}".format(module, title) if module and module != title else title
    steps = "\n".join(step_parts).strip() if step_parts else "-"
    expected_result = "\n".join(expected_parts).strip() if expected_parts else "-"
    raw_text = " ".join([display_title, steps, expected_result]).strip()
    return ParsedTestCase(
        title=display_title,
        steps=steps,
        expected_result=expected_result,
        raw_text=raw_text,
    )
